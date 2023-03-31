import csv
import zipfile

import openpyxl
import pytest
import requests
import os
from os.path import abspath, dirname, join

from selene import browser, be, have
from selenium import webdriver
from selenium.webdriver.chrome.service import Service as ChromeService
from webdriver_manager.chrome import ChromeDriverManager

from openpyxl import load_workbook
from PyPDF2 import PdfReader

from zipfile import ZipFile

import io


XLSX_DOWNLOAD_LINK = 'https://freetestdata.com/document-files/xlsx/'  # [role="button" ][href*="100KB_XLSX"]
PDF_DOWNLOAD_LINK = 'https://laserscom.com/ru/products'  # [role="button"][href*="100KB_PDF"]
CSV_DOWNLOAD_LINK = 'https://freetestdata.com/wp-content/uploads/2021/09/Free_Test_Data_200KB_CSV-1.csv'  # [role="button" ][href*="200KB_CSV"]

BASE_DIR = dirname(dirname(abspath(__file__)))  # path to project root directory
RESOURCES_DIR = join(BASE_DIR, 'resources')  # path to save files

CSV_TEST_FILE_NAME = 'csv_test_file.csv'
ZIP_FILE_NAME =     'archive.zip'


@pytest.fixture(scope='function', autouse=True)
def create_resources_for_tests():
    if not os.path.exists(RESOURCES_DIR):
        os.makedirs(RESOURCES_DIR)

    # setup selene browser for files download
    options = webdriver.ChromeOptions()  # создаем объект options на базе вэбдрайвера Selenium

    # dictionary that specify various preferences for the webdriver.
    prefs = {
        "download.default_directory": RESOURCES_DIR,
        "plugins.always_open_pdf_externally": True,
        "download.prompt_for_download": False
    }
    options.add_experimental_option("prefs", prefs)  # передаем в options значения из словаря prefs

    # create a new Chrome webdriver instance with specific options and a service object.
    # the resulting driver object can be used to control and automate the Chrome browser
    driver = webdriver.Chrome(service=ChromeService(ChromeDriverManager().install()), options=options)

    browser.config.driver = driver  # настраиваем браузер selene на использование созданного драйвера
    browser.config.hold_browser_open = True

    browser.open(PDF_DOWNLOAD_LINK)
    browser.all('.views-field-title .field-content a').element_by(have.exact_text('LDS-505-FP-10')) \
         .should(be.clickable).click()

    browser.open(XLSX_DOWNLOAD_LINK)
    browser.element('[role="button" ][href*="100KB_XLSX"]').should(be.clickable).click()

    # скачиваем файл используя библиотеку requesta
    response = requests.get(CSV_DOWNLOAD_LINK, allow_redirects=True)
    csv_file_path = join(RESOURCES_DIR, CSV_TEST_FILE_NAME)
    with open(csv_file_path, 'wb') as csv_test_file:
        csv_test_file.write(response.content)

    yield
    # TODO раскоментировать
    # browser.close()
    # browser.quit()





def test_check_files_in_zip_archive():
    # определяем число скачанных файлов
    downloaded_files_list = [f for f in os.listdir(RESOURCES_DIR)]
    downloaded_files_number = len(downloaded_files_list)

    # записываем имена скачанных файлов в переменные
    downloaded_pdf_file_name = "".join([f for f in downloaded_files_list if f.endswith('.pdf')])
    downloaded_xlsx_file_name = "".join([f for f in downloaded_files_list if f.endswith('.xlsx')])

    # определяем размер скачанных файлов файлов
    downloaded_pdf_file_size = os.path.getsize(join(RESOURCES_DIR, downloaded_pdf_file_name))
    downloaded_csv_file_size = os.path.getsize(join(RESOURCES_DIR, CSV_TEST_FILE_NAME))
    downloaded_xlsx_file_size = os.path.getsize(join(RESOURCES_DIR, downloaded_xlsx_file_name))
    downloaded_files_size_list = [downloaded_pdf_file_size, downloaded_csv_file_size, downloaded_xlsx_file_size]

    # cоздаем для загруженных файлов контрольные данные
    # XLSX файл
    """
    Для XLSX:
    - downloaded_xlsx_file_sheet_names - список листов книги 
    - downloaded_xlsx_file_first_sheet_name - название первого листа 
    - downloaded_xlsx_file_first_sheet_max_col - количество столбцов в перовом листе
    - downloaded_xlsx_file_first_sheet_max_row - количество строк в первом листе 
    - downloaded_xlsx_file_first_sheet_first_row_value  - содержимое первой строки первой строки 
    - downloaded_xlsx_file_first_sheet_last_column_value - содержимое последнего столбца
    - downloaded_xlsx_file_first_sheet_b2_sell - содержимое ячейки B2
    """
    downloaded_xlsx_file_path = join(RESOURCES_DIR, downloaded_xlsx_file_name)
    downloaded_xlsx_file = openpyxl.load_workbook(join(RESOURCES_DIR, downloaded_xlsx_file_name))
    downloaded_xlsx_file_first_sheet = downloaded_xlsx_file.worksheets[0]

    downloaded_xlsx_file_sheet_names = downloaded_xlsx_file.sheetnames
    downloaded_xlsx_file_first_sheet_max_col = downloaded_xlsx_file_first_sheet.max_column
    downloaded_xlsx_file_first_sheet_max_row = downloaded_xlsx_file_first_sheet.max_row
    downloaded_xlsx_file_first_sheet_b2_sell = downloaded_xlsx_file_first_sheet.cell(row = 2, column = 2).value

    def get_xlsx_row_values(path, row):
        wb = openpyxl.load_workbook(path)
        first_sheet = wb.worksheets[0]
        row_values = [first_sheet.cell(row=row, column=col).value for col in range(1, first_sheet.max_column + 1)]
        return row_values

    downloaded_xlsx_file_first_sheet_first_row_value = get_xlsx_row_values(join(RESOURCES_DIR, downloaded_xlsx_file_name), 1)

    def get_xlsx_column_values(path, column):
        wb = openpyxl.load_workbook(path)
        first_sheet = wb.worksheets[0]
        column_values = [first_sheet.cell(column=column, row=r).value for r in
                         range(1, first_sheet.max_row + 1)]
        return column_values

    downloaded_xlsx_file_first_sheet_last_column_value = get_xlsx_column_values\
        (join(RESOURCES_DIR, downloaded_xlsx_file_name), downloaded_xlsx_file_first_sheet_max_col)

    # PDF файл
    downloaded_pdf_file_path = join(RESOURCES_DIR, downloaded_pdf_file_name)
    with open(downloaded_pdf_file_path, 'rb') as pdf_file:
        pdf_file_reader = PdfReader(pdf_file)
        downloaded_pdf_file_number_pages = len(pdf_file_reader.pages)
        downloaded_pdf_file_first_page_text = pdf_file_reader.pages[0].extract_text()
        downloaded_pdf_file_last_page_text = pdf_file_reader.pages[downloaded_pdf_file_number_pages - 1].extract_text()
        downloaded_pdf_file_meta_title = pdf_file_reader.metadata.title

    # CSV файл
    downloaded_csv_file_path = join(RESOURCES_DIR,  CSV_TEST_FILE_NAME)

    def get_csv_row_value(csv_file_path, row_num):
        with open(downloaded_csv_file_path, 'rt') as csv_file:
            csv_file_reader = csv.reader(csv_file)
            for i, row in enumerate(csv_file_reader):
                if i == row_num - 1:
                    row_value = row
        return row_value
    downloaded_csv_file_row_2_value = get_csv_row_value(downloaded_csv_file_path, 2)

    def get_csv_col_value(csv_file_path, col_name):
        with open(csv_file_path, 'rt') as csv_file:
            csv_file_reader = csv.DictReader(csv_file)
            col_value = []
            for row in csv_file_reader:
                if row['COUNTRY'] != '':
                    col_value.append(row[col_name])
        return col_value

    downloaded_csv_file_col_name_value = get_csv_col_value(downloaded_csv_file_path, 'NAME')

    # zip downloaded files to archive.zip file in resouces directory
    zip_file_path = join(RESOURCES_DIR, ZIP_FILE_NAME)
    zip_files_list = [downloaded_csv_file_path, downloaded_pdf_file_path, downloaded_xlsx_file_path]
    with zipfile.ZipFile(zip_file_path, mode='w') as zip_file_object:
        for file in zip_files_list:
            zip_file_object.write(file)

    # get check data from zip file
    with zipfile.ZipFile(zip_file_path, 'r') as zip_file_object:
        # Получаем список имен файлов в архиве
        file_names = zip_file_object.namelist()

        zip_csv_file_path = ''.join(filter(lambda x: '.csv' in x, file_names))
        zip_pdf_file_path = ''.join(filter(lambda x: '.pdf' in x, file_names))
        zip_xlsx_file_path = ''.join(filter(lambda x: '.xlsx' in x, file_names))

        # создаем список размеров файлов в архиве
        zip_files_size_list = [zip_file_object.getinfo(file_name).file_size for file_name in file_names]

        # читаем данные из CSV-файла в архиве
        with zip_file_object.open(zip_csv_file_path, 'r') as csv_file_object:
            # csv_file_object_reader = csv.reader(csv_file_object)
            csv_file_object_reader = csv.reader(io.StringIO(csv_file_object.read().decode('utf-8')))
            for i, row in enumerate(csv_file_object_reader):
                if i == 2-1:
                    zip_csv_file_row_2_value = row
                    break

        # читаем данные из PDF-файла в архиве
        with zip_file_object.open(zip_pdf_file_path, 'r') as pdf_file_object:
            pdf_file_reader = PdfReader(pdf_file_object)
            zip_pdf_file_number_pages = len(pdf_file_reader.pages)
            zip_pdf_file_first_page_text = pdf_file_reader.pages[0].extract_text()
            zip_pdf_file_last_page_text = pdf_file_reader.pages[
            zip_pdf_file_number_pages - 1].extract_text()
            zip_pdf_file_meta_title = pdf_file_reader.metadata.title

        # читаем данные из XLSX-файла в архиве
        with zip_file_object.open(zip_xlsx_file_path, 'r') as xlsx_file_object:
            zip_xlsx_file = openpyxl.load_workbook(xlsx_file_object)
            zip_xlsx_file_first_sheet = zip_xlsx_file.worksheets[0]

            zip_xlsx_file_sheet_names = zip_xlsx_file.sheetnames
            zip_xlsx_file_first_sheet_max_col = zip_xlsx_file_first_sheet.max_column
            zip_xlsx_file_first_sheet_max_row = zip_xlsx_file_first_sheet.max_row
            zip_xlsx_file_first_sheet_b2_sell = zip_xlsx_file_first_sheet.cell(row=2, column=2).value

    assert zip_files_size_list.sort() == downloaded_files_size_list.sort()

    assert zip_csv_file_row_2_value == downloaded_csv_file_row_2_value

    assert zip_pdf_file_number_pages == downloaded_pdf_file_number_pages
    assert zip_pdf_file_last_page_text == downloaded_pdf_file_last_page_text
    assert zip_pdf_file_number_pages == downloaded_pdf_file_number_pages
    assert zip_pdf_file_meta_title == downloaded_pdf_file_meta_title




    # os.remove(join(RESOURCES_DIR, CSV_TEST_FILE_NAME))
    # os.remove(join(RESOURCES_DIR, downloaded_pdf_file_name))
    # os.remove(join(RESOURCES_DIR, downloaded_xlsx_file_name))
    # os.rmdir(RESOURCES_DIR)
